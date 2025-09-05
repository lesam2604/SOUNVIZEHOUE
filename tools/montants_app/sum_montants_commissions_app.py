import os
import re
import sys
import csv
from decimal import Decimal, InvalidOperation, getcontext
from dataclasses import dataclass
from typing import Optional, Tuple

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

# High precision for money
getcontext().prec = 28


# Optional deps (loaded lazily when needed)
try:
    from openpyxl import load_workbook  # type: ignore
except Exception:
    load_workbook = None  # type: ignore

try:
    from pdfminer.high_level import extract_text as pdf_extract_text  # type: ignore
except Exception:
    pdf_extract_text = None  # type: ignore

try:
    from reportlab.lib.pagesizes import A4  # type: ignore
    from reportlab.pdfgen import canvas  # type: ignore
except Exception:
    A4 = None  # type: ignore
    canvas = None  # type: ignore


NBSP = "\u00A0"

# Numbers like: 1 234,56 or 1234,56 or -14 100,00 or 500,01 or 250,00
NUM_RE = re.compile(
    rf"""
    (?<!\w)
    \(?
    -?
    (?:
        \d{{1,3}}(?:[ .,{NBSP}]\d{{3}})+  # thousands separated
        |
        \d+
    )
    (?:[.,]\d+)?
    \)?
    (?!\w)
    """,
    re.VERBOSE,
)

# Be permissive about date separators (e.g., '-', 'â€‘', 'â€“') to support PDF extracts
DATE_TIME_RE = re.compile(r"\b\d{2}\D\d{2}\D\d{2}\s+\d{2}:\d{2}:\d{2}\b")


def normalize_number(token: str) -> Optional[Decimal]:
    if not token:
        return None
    s = token.strip()

    neg = s.startswith("(") and s.endswith(")")
    s = s.strip("()")

    s = s.replace(" ", "").replace(NBSP, "")

    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")

    s = re.sub(r"[^0-9\.-]", "", s)
    if s in {"", ".", "-", "-."}:
        return None
    try:
        val = Decimal(s)
        if neg:
            val = -val
        return val
    except (InvalidOperation, ValueError):
        return None


def pretty_amount(d: Decimal) -> str:
    s = f"{d:,.2f}"
    s = s.replace(",", " ").replace(".", ",")
    return s


@dataclass
class Totals:
    amount_total: Decimal
    commission_total: Decimal
    recharge_total: Decimal
    rows_count: int
    commission_rows_count: int
    recharge_rows_count: int
    source_path: str


@dataclass
class ParsedRow:
    dt_text: str
    amount: Decimal
    description: str


def is_commission_text(text: str) -> bool:
    t = text.lower()
    # Strict rule: only lines whose Description contains 'commission'
    return "commission" in t


def is_recharge_text(text: str) -> bool:
    # Strict: match the full word 'Recharge' only (case-insensitive)
    return re.search(r"\brecharge\b", text, flags=re.IGNORECASE) is not None


def parse_text_report(text: str, source_path: str) -> Totals:
    # We will only count amounts for lines classified as Commission or Recharge
    commission_total = Decimal(0)
    recharge_total = Decimal(0)
    rows = 0
    commission_rows = 0
    recharge_rows = 0

    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue

        # Skip headers
        if line.startswith("Distribution Account Activity Report"):
            continue
        if line.startswith("Trans ID "):
            continue

        m = DATE_TIME_RE.search(line)
        if not m:
            continue

        after_dt = line[m.end():]
        nums = [normalize_number(n.group(0)) for n in NUM_RE.finditer(after_dt)]
        nums = [n for n in nums if n is not None]
        if not nums:
            continue

        amount = nums[0]
        rows += 1
        is_comm = is_commission_text(after_dt)
        is_rech = is_recharge_text(after_dt)
        if is_comm:
            commission_total += amount
            commission_rows += 1
        if is_rech:
            recharge_total += amount
            recharge_rows += 1

    # Total Montants is the sum of the two categories only
    amount_total = commission_total + recharge_total

    return Totals(
        amount_total=amount_total,
        commission_total=commission_total,
        recharge_total=recharge_total,
        rows_count=rows,
        commission_rows_count=commission_rows,
        recharge_rows_count=recharge_rows,
        source_path=source_path,
    )


def parse_text_to_rows(text: str) -> list[ParsedRow]:
    rows: list[ParsedRow] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("Distribution Account Activity Report"):
            continue
        if line.startswith("Trans ID "):
            continue
        m = DATE_TIME_RE.search(line)
        if not m:
            continue
        dt_text = m.group(0)
        after_dt = line[m.end():]
        nums = [normalize_number(n.group(0)) for n in NUM_RE.finditer(after_dt)]
        nums = [n for n in nums if n is not None]
        if not nums:
            continue
        amount = nums[0]
        desc = after_dt.strip()
        rows.append(ParsedRow(dt_text=dt_text, amount=amount, description=desc))
    return rows


def totals_from_rows(rows: list[ParsedRow]) -> Totals:
    commission_total = Decimal(0)
    recharge_total = Decimal(0)
    commission_rows = 0
    recharge_rows = 0
    for r in rows:
        is_comm = is_commission_text(r.description)
        is_rech = is_recharge_text(r.description)
        if is_comm:
            commission_total += r.amount
            commission_rows += 1
        if is_rech:
            recharge_total += r.amount
            recharge_rows += 1
    amount_total = commission_total + recharge_total
    return Totals(
        amount_total=amount_total,
        commission_total=commission_total,
        recharge_total=recharge_total,
        rows_count=len(rows),
        commission_rows_count=commission_rows,
        recharge_rows_count=recharge_rows,
        source_path="(texte collÃ©)",
    )


def sum_text_file(path: str) -> Totals:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(path, "r", encoding=enc, errors="ignore") as f:
                text = f.read()
            return parse_text_report(text, path)
        except Exception:
            continue
    return parse_text_report("", path)


def sum_pdf(path: str) -> Totals:
    if not pdf_extract_text:
        raise RuntimeError("pdfminer.six n'est pas installÃ©. pip install pdfminer.six")
    text = pdf_extract_text(path) or ""
    return parse_text_report(text, path)


def sum_csv(path: str) -> Totals:
    # Only count rows whose Description is Commission or Recharge
    commission_total = Decimal(0)
    recharge_total = Decimal(0)
    rows = 0
    commission_rows = 0
    recharge_rows = 0

    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(path, "r", encoding=enc, newline="", errors="ignore") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
                except Exception:
                    dialect = csv.excel

                reader = csv.DictReader(f, dialect=dialect)
                headers = [h or "" for h in (reader.fieldnames or [])]
                low_headers = [h.lower() for h in headers]

                # Try to locate columns
                def find_col(*cands):
                    for c in cands:
                        c_low = c.lower()
                        if c_low in low_headers:
                            return headers[low_headers.index(c_low)]
                    return None

                amount_col = find_col("Amount", "Montant")
                desc_col = find_col("Description", "Libelle", "Label")

                if amount_col:
                    for row in reader:
                        amount = normalize_number(str(row.get(amount_col, "")))
                        if amount is None:
                            continue
                        rows += 1
                        desc_text = str(row.get(desc_col, "")) if desc_col else ""
                        is_comm = is_commission_text(desc_text)
                        is_rech = is_recharge_text(desc_text)
                        if is_comm:
                            commission_total += amount
                            commission_rows += 1
                        if is_rech:
                            recharge_total += amount
                            recharge_rows += 1
                else:
                    # Fallback: scan row cells, try to infer amount as the first monetary token after date/time pattern
                    f.seek(0)
                    raw_reader = csv.reader(f, dialect=dialect)
                    for row in raw_reader:
                        line = " ".join(str(c) for c in row)
                        m = DATE_TIME_RE.search(line)
                        if not m:
                            continue
                        after_dt = line[m.end():]
                        nums = [normalize_number(n.group(0)) for n in NUM_RE.finditer(after_dt)]
                        nums = [n for n in nums if n is not None]
                        if not nums:
                            continue
                        amount = nums[0]
                        rows += 1
                        is_comm = is_commission_text(after_dt)
                        is_rech = is_recharge_text(after_dt)
                        if is_comm:
                            commission_total += amount
                            commission_rows += 1
                        if is_rech:
                            recharge_total += amount
                            recharge_rows += 1

                break
        except Exception:
            continue

    # Total Montants is the sum of the two categories only
    amount_total = commission_total + recharge_total

    return Totals(
        amount_total=amount_total,
        commission_total=commission_total,
        recharge_total=recharge_total,
        rows_count=rows,
        commission_rows_count=commission_rows,
        recharge_rows_count=recharge_rows,
        source_path=path,
    )


def sum_excel(path: str) -> Totals:
    if not load_workbook:
        raise RuntimeError("openpyxl n'est pas installÃ©. pip install openpyxl")

    commission_total = Decimal(0)
    recharge_total = Decimal(0)
    rows = 0
    commission_rows = 0
    recharge_rows = 0

    wb = load_workbook(filename=path, data_only=True, read_only=True)
    for ws in wb.worksheets:
        iter_rows = ws.iter_rows(values_only=True)
        try:
            first = next(iter_rows)
        except StopIteration:
            continue

        headers = None
        if first and any(isinstance(x, str) for x in first):
            headers = [str(x) if x is not None else "" for x in first]

        if headers:
            low_headers = [h.lower() for h in headers]
            def find_idx(*cands):
                for c in cands:
                    c_low = c.lower()
                    if c_low in low_headers:
                        return low_headers.index(c_low)
                return None

            idx_amount = find_idx("Amount", "Montant")
            idx_desc = find_idx("Description", "Libelle", "Label")

            if idx_amount is not None:
                for row in iter_rows:
                    if not row or idx_amount >= len(row):
                        continue
                    cell = row[idx_amount]
                    if isinstance(cell, (int, float, Decimal)):
                        amount = Decimal(str(cell))
                    else:
                        amount = normalize_number(str(cell))
                    if amount is None:
                        continue
                    rows += 1
                    desc_text = str(row[idx_desc]) if (idx_desc is not None and idx_desc < len(row)) else ""
                    is_comm = is_commission_text(desc_text)
                    is_rech = is_recharge_text(desc_text)
                    if is_comm:
                        commission_total += amount
                        commission_rows += 1
                    if is_rech:
                        recharge_total += amount
                        recharge_rows += 1
                continue

        # Fallback: scan by lines composed from rows
        text_lines = []
        # Include header row if present
        text_lines.append(" ".join(str(x) for x in (first or [])))
        for row in ws.iter_rows(values_only=True):
            text_lines.append(" ".join("" if v is None else str(v) for v in row))
        totals = parse_text_report("\n".join(text_lines), path)
        # The fallback text parser already applies strict classification
        commission_total += totals.commission_total
        recharge_total += totals.recharge_total
        rows += totals.rows_count
        commission_rows += totals.commission_rows_count
        recharge_rows += totals.recharge_rows_count

    # Total Montants is the sum of the two categories only
    amount_total = commission_total + recharge_total

    return Totals(
        amount_total=amount_total,
        commission_total=commission_total,
        recharge_total=recharge_total,
        rows_count=rows,
        commission_rows_count=commission_rows,
        recharge_rows_count=recharge_rows,
        source_path=path,
    )


def compute_totals_for_file(path: str) -> Totals:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv",):
        return sum_csv(path)
    if ext in (".xlsx", ".xls"):
        return sum_excel(path)
    if ext in (".pdf",):
        return sum_pdf(path)
    # default: treat as text export
    return sum_text_file(path)


def export_csv(totals: Totals, out_path: str) -> None:
    with open(out_path, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["Source", totals.source_path])
        w.writerow(["Lignes traitÃ©es", totals.rows_count])
        w.writerow([])
        w.writerow(["Categorie", "Total"])
        w.writerow(["Total Montants (tous)", str(totals.amount_total)])
        w.writerow(["Total Commissions", str(totals.commission_total)])
        w.writerow(["Total Hors Commission (Recharges)", str(totals.recharge_total)])


def export_pdf(totals: Totals, out_path: str) -> None:
    if not (A4 and canvas):
        raise RuntimeError("reportlab n'est pas installÃ©. pip install reportlab")
    c = canvas.Canvas(out_path, pagesize=A4)
    width, height = A4
    x = 50
    y = height - 50
    line_h = 18

    c.setFont("Helvetica-Bold", 14)
    c.drawString(x, y, "Somme des Montants et Commissions")
    y -= 2 * line_h
    c.setFont("Helvetica", 10)
    c.drawString(x, y, f"Source: {totals.source_path}")
    y -= line_h
    c.drawString(x, y, f"Lignes traitÃ©es: {totals.rows_count}")
    y -= 2 * line_h

    c.setFont("Helvetica-Bold", 12)
    c.drawString(x, y, "RÃ©sultats")
    y -= line_h
    c.setFont("Helvetica", 11)
    c.drawString(x, y, f"Total Montants (tous): {pretty_amount(totals.amount_total)}")
    y -= line_h
    c.drawString(x, y, f"Total Commissions: {pretty_amount(totals.commission_total)}")
    y -= line_h
    c.drawString(x, y, f"Total Hors Commission (Recharges): {pretty_amount(totals.recharge_total)}")

    c.showPage()
    c.save()


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Somme Montants & Commissions")
        self.geometry("680x360")

        # ttk theming
        try:
            style = ttk.Style(self)
            style.theme_use("clam")
        except Exception:
            pass

        self.totals: Optional[Totals] = None
        self.parsed_rows: list[ParsedRow] = []

        # Menu
        menubar = tk.Menu(self)
        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Ouvrir (Ctrl+O)", command=self.on_open)
        filemenu.add_separator()
        filemenu.add_command(label="Exporter CSV (Ctrl+S)", command=self.on_export_csv, state=tk.DISABLED)
        filemenu.add_command(label="Exporter PDF (Ctrl+P)", command=self.on_export_pdf, state=tk.DISABLED)
        filemenu.add_separator()
        filemenu.add_command(label="Quitter", command=self.destroy)
        menubar.add_cascade(label="Fichier", menu=filemenu)
        self.config(menu=menubar)

        # Toolbar
        toolbar = ttk.Frame(self, padding=(10,8))
        toolbar.pack(fill=tk.X)
        self.btn_open = ttk.Button(toolbar, text="Ouvrir", command=self.on_open)
        self.btn_open.pack(side=tk.LEFT)
        self.btn_csv = ttk.Button(toolbar, text="Exporter CSV", command=self.on_export_csv, state=tk.DISABLED)
        self.btn_csv.pack(side=tk.LEFT, padx=(8,0))
        self.btn_pdf = ttk.Button(toolbar, text="Exporter PDF", command=self.on_export_pdf, state=tk.DISABLED)
        self.btn_pdf.pack(side=tk.LEFT, padx=(8,0))
        self.btn_paste = ttk.Button(toolbar, text="Coller texte", command=self.on_paste_text)
        self.btn_paste.pack(side=tk.LEFT, padx=(12,0))

        # Content
        content = ttk.Frame(self, padding=12)
        content.pack(fill=tk.BOTH, expand=True)

        # Cards area
        cards = ttk.Frame(content)
        cards.pack(fill=tk.X, pady=(0,10))

        def make_card(parent, bg, title):
            frame = tk.Frame(parent, bg=bg, bd=0, highlightthickness=0, padx=14, pady=10)
            title_lbl = tk.Label(frame, text=title, bg=bg, fg="white", font=("Segoe UI", 9))
            value_lbl = tk.Label(frame, text="â€”", bg=bg, fg="white", font=("Segoe UI", 16, "bold"))
            title_lbl.pack(anchor="w")
            value_lbl.pack(anchor="w")
            return frame, value_lbl

        self.card_amount, self.val_amount = make_card(cards, "#2563eb", "Total Montants")   # blue-600
        self.card_comm, self.val_comm   = make_card(cards, "#16a34a", "Total Commissions")  # green-600
        self.card_non, self.val_non     = make_card(cards, "#334155", "Hors Commission")    # slate-700

        self.card_amount.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.card_comm.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)
        self.card_non.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Details area
        details = ttk.Frame(content)
        details.pack(fill=tk.BOTH, expand=True)
        self.lbl_source = ttk.Label(details, text="Source: -")
        self.lbl_source.pack(anchor="w")
        self.lbl_rows = ttk.Label(details, text="Lignes traitees: -   |   Lignes commission: -   |   Lignes recharge: -")
        self.lbl_rows.pack(anchor="w")

        # Table preview for parsed text
        table_frame = ttk.Frame(details)
        table_frame.pack(fill=tk.BOTH, expand=True, pady=(8,0))
        cols = ("date", "amount", "description")
        self.table = ttk.Treeview(table_frame, columns=cols, show="headings", height=8)
        self.table.heading("date", text="Date/Heure")
        self.table.heading("amount", text="Amount")
        self.table.heading("description", text="Description")
        self.table.column("date", width=120, anchor="w")
        self.table.column("amount", width=100, anchor="e")
        self.table.column("description", width=400, anchor="w")
        yscroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.table.yview)
        self.table.configure(yscrollcommand=yscroll.set)
        self.table.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        yscroll.pack(side=tk.RIGHT, fill=tk.Y)

        # Status bar
        self.status = tk.StringVar(value="PrÃªt")
        statusbar = ttk.Label(self, textvariable=self.status, padding=(10,4), anchor="w")
        statusbar.pack(fill=tk.X)

        # Shortcuts
        self.bind_all("<Control-o>", lambda e: self.on_open())
        self.bind_all("<Control-s>", lambda e: self.on_export_csv())
        self.bind_all("<Control-p>", lambda e: self.on_export_pdf())

    def on_open(self) -> None:
        path = filedialog.askopenfilename(
            title="SÃ©lectionner le fichier du rapport",
            filetypes=[
                ("Tous fichiers", "*.*"),
                ("Excel", "*.xlsx;*.xls"),
                ("CSV", "*.csv"),
                ("PDF", "*.pdf"),
                ("Texte", "*.txt;*.log"),
            ],
        )
        if not path:
            return
        try:
            self.totals = compute_totals_for_file(path)
            self.refresh_view()
            self.btn_csv.config(state=tk.NORMAL)
            self.btn_pdf.config(state=tk.NORMAL)
            # menu enabling removed to avoid Tk 'menu' widget lookup errors
            self.status.set("Fichier chargÃ©")
        except Exception as e:
            messagebox.showerror("Erreur", str(e))
            self.status.set("Erreur: " + str(e))

    def refresh_view(self) -> None:
        if not self.totals:
            return
        t = self.totals
        self.lbl_source.config(text=f"Source: {t.source_path}")
        self.lbl_rows.config(text=f"Lignes traitees: -   |   Lignes commission: -   |   Lignes recharge: -")
        self.val_amount.config(text=pretty_amount(t.amount_total))
        self.val_comm.config(text=pretty_amount(t.commission_total))
        self.val_non.config(text=pretty_amount(t.recharge_total))

        # Populate table if available
        if hasattr(self, 'table'):
            self.table.delete(*self.table.get_children())
            for r in self.parsed_rows[:1000]:
                self.table.insert("", tk.END, values=(r.dt_text, pretty_amount(r.amount), r.description[:200]))

    def on_export_csv(self) -> None:
        if not self.totals:
            return
        default = os.path.join(
            os.path.dirname(self.totals.source_path),
            "resultats_montants_commissions.csv",
        )
        out = filedialog.asksaveasfilename(
            title="Enregistrer le CSV",
            defaultextension=".csv",
            initialfile=os.path.basename(default),
            filetypes=[("CSV", "*.csv")],
        )
        if not out:
            return
        try:
            export_csv(self.totals, out)
            messagebox.showinfo("Export CSV", f"CSV crÃ©Ã©:\n{out}")
            self.status.set(f"CSV exportÃ©: {out}")
        except Exception as e:
            messagebox.showerror("Erreur", str(e))
            self.status.set("Erreur export CSV: " + str(e))

    def on_export_pdf(self) -> None:
        if not self.totals:
            return
        default = os.path.join(
            os.path.dirname(self.totals.source_path),
            "resultats_montants_commissions.pdf",
        )
        out = filedialog.asksaveasfilename(
            title="Enregistrer le PDF",
            defaultextension=".pdf",
            initialfile=os.path.basename(default),
            filetypes=[("PDF", "*.pdf")],
        )
        if not out:
            return
        try:
            export_pdf(self.totals, out)
            messagebox.showinfo("Export PDF", f"PDF crÃ©Ã©:\n{out}")
            self.status.set(f"PDF exportÃ©: {out}")
        except Exception as e:
            messagebox.showerror("Erreur", str(e))
            self.status.set("Erreur export PDF: " + str(e))

    def on_paste_text(self) -> None:
        dlg = tk.Toplevel(self)
        dlg.title("Coller du texte brut")
        dlg.geometry("720x420")
        txt = tk.Text(dlg, wrap="word")
        txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        btns = ttk.Frame(dlg)
        btns.pack(fill=tk.X, padx=8, pady=(0,8))

        def analyze():
            content = txt.get("1.0", tk.END)
            rows = parse_text_to_rows(content)
            self.parsed_rows = rows
            self.totals = totals_from_rows(rows)
            self.btn_csv.config(state=tk.NORMAL)
            self.btn_pdf.config(state=tk.NORMAL)
            self.refresh_view()
            self.status.set(f"Texte analysé: {len(rows)} lignes")

        ttk.Button(btns, text="Analyser", command=analyze).pack(side=tk.LEFT)
        ttk.Button(btns, text="Fermer", command=dlg.destroy).pack(side=tk.LEFT, padx=(8,0))


def main() -> None:
    # If run with a file path, do a quick CLI summary
    if len(sys.argv) > 1:
        path = sys.argv[1]
        totals = compute_totals_for_file(path)
        print(f"Source: {totals.source_path}")
        print(f"Lignes: {totals.rows_count}")
        print(f"Total Montants (tous): {pretty_amount(totals.amount_total)}")
        print(f"Total Commissions: {pretty_amount(totals.commission_total)}")
        print(f"Total Hors Commission (Recharges): {pretty_amount(totals.recharge_total)}")
        return

    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()


